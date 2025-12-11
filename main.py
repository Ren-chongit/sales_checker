import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import pdfplumber
import re
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def select_file(title, filetypes):
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    return file_path

def extract_pdf_data(pdf_path):
    data = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Check if page contains the target header
            if "科目別合計" not in page.extract_text():
                continue
                
            words = page.extract_words()
            # Group words by line (using 'top' coordinate)
            lines = {}
            for word in words:
                top = word['top']
                # Find existing line with similar top (tolerance of 3)
                found_line = False
                for line_top in lines:
                    if abs(line_top - top) < 3:
                        lines[line_top].append(word)
                        found_line = True
                        break
                if not found_line:
                    lines[top] = [word]
            
            # Sort lines by top (vertical order)
            sorted_line_tops = sorted(lines.keys())
            
            for top in sorted_line_tops:
                line_words = sorted(lines[top], key=lambda w: w['x0'])
                
                # Filter out header lines
                line_text = "".join([w['text'] for w in line_words])
                if "科目別合計" in line_text or "仕入先" in line_text or "原価" in line_text:
                    continue
                
                # Parse line words
                # Expected pattern: Code -> Name (1+ words) -> Cost -> Com -> Repeat
                
                idx = 0
                while idx < len(line_words):
                    # 1. Identify Code
                    # Code is usually alphanumeric.
                    # If we are at the end, break
                    if idx >= len(line_words):
                        break
                        
                    code_word = line_words[idx]
                    code = code_word['text']
                    idx += 1
                    
                    # 2. Identify Name and Cost
                    # Name can be multiple words. Cost is a number (digits and commas).
                    # Cost must be followed by Com (another number).
                    
                    name_parts = []
                    cost = None
                    
                    while idx < len(line_words):
                        word = line_words[idx]
                        text = word['text']
                        
                        # Check if this word is a potential Cost
                        # It must be a number, AND the NEXT word must also be a number (Com)
                        is_number = re.match(r'^[\d,]+$', text)
                        
                        is_cost = False
                        if is_number:
                            # Check next word
                            if idx + 1 < len(line_words):
                                next_text = line_words[idx+1]['text']
                                if re.match(r'^[\d,]+$', next_text):
                                    is_cost = True
                        
                        if is_cost:
                            cost = int(text.replace(',', ''))
                            idx += 1 # Consume Cost
                            
                            # Consume Com
                            if idx < len(line_words):
                                # com = int(line_words[idx]['text'].replace(',', '')) # We don't need Com but we consume it
                                idx += 1
                            break
                        else:
                            name_parts.append(text)
                            idx += 1
                    
                    if cost is not None:
                        name = "".join(name_parts)
                        data.append({
                            'pdf_code': code,
                            'pdf_name': name,
                            'pdf_cost': cost
                        })
                    else:
                        # If we didn't find a cost, this "Code" might have been garbage or end of line
                        pass
    return pd.DataFrame(data)

def main():
    # 1. Select Cost Data File (CSV)
    csv_path = select_file("1. 原価データファイルを選択してください (CSV)", [("CSV Files", "*.csv")])
    if not csv_path:
        messagebox.showinfo("キャンセル", "ファイル選択がキャンセルされました。")
        return

    # 2. Select Transferred Data File (PDF)
    pdf_path = select_file("2. 振替済みデータファイルを選択してください (PDF)", [("PDF Files", "*.pdf")])
    if not pdf_path:
        messagebox.showinfo("キャンセル", "ファイル選択がキャンセルされました。")
        return

    try:
        # Read CSV
        # Assuming CSV has headers. Based on `head` output: コード,仕入先名,合計原価,...
        df_csv = pd.read_csv(csv_path, encoding='cp932')
        
        # Extract PDF Data
        df_pdf = extract_pdf_data(pdf_path)
        
        if df_pdf.empty:
            messagebox.showwarning("警告", "PDFからデータを抽出できませんでした。「科目別合計」のページが見つからないか、形式が異なります。")
            return

        # Merge Data
        # Left join on Code to keep all CSV records, and Outer join to find unmatched PDF records?
        # Requirement: "csvのコードとpdfの仕入先（数字とアルファベット）がマッチするものを合体"
        # Requirement: "csvのコードとpdfの仕入先（数字とアルファベット）がマッチしないもの...次に記載"
        
        # Let's use an outer join to get all records from both
        merged = pd.merge(df_csv, df_pdf, left_on='コード', right_on='pdf_code', how='outer')
        
        # Prepare Output Columns
        # "コード", "仕入れ先名"（名称は原価データを優先）, "原価一覧の原価", "振替済みの原価"
        
        output_data = []
        
        # Separate into categories for sorting/ordering
        mismatched_cost = []
        unmatched_code = []
        matched = []
        
        for index, row in merged.iterrows():
            code = row['コード'] if pd.notna(row['コード']) else row['pdf_code']
            name = row['仕入先名'] if pd.notna(row['仕入先名']) else row['pdf_name']
            csv_cost = row['合計原価'] if pd.notna(row['合計原価']) else 0
            pdf_cost = row['pdf_cost'] if pd.notna(row['pdf_cost']) else 0
            
            # Ensure costs are integers for comparison
            try:
                csv_cost = int(csv_cost)
            except:
                csv_cost = 0
            
            try:
                pdf_cost = int(pdf_cost)
            except:
                pdf_cost = 0

            record = {
                "コード": code,
                "仕入れ先名": name,
                "原価一覧の原価": csv_cost,
                "振替済みの原価": pdf_cost,
                "_status": "" # internal helper
            }
            
            if pd.isna(row['コード']) or pd.isna(row['pdf_code']):
                # Code not matching (exists in one but not other)
                record["_status"] = "unmatched_code"
                unmatched_code.append(record)
            elif csv_cost != pdf_cost:
                # Code matches but cost mismatch
                record["_status"] = "mismatched_cost"
                mismatched_cost.append(record)
            else:
                # Fully matched
                record["_status"] = "matched"
                matched.append(record)
                
        # Combine in order: Mismatched Cost -> Unmatched Code -> Matched -> Others (if any logic for others)
        final_rows = mismatched_cost + unmatched_code + matched
        
        # Create Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "マッチング結果"
        
        headers = ["コード", "仕入れ先名", "原価一覧の原価", "振替済みの原価"]
        ws.append(headers)
        
        # Fills
        fill_yellow = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid") # Light Yellow
        fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")    # Light Red
        
        for row_data in final_rows:
            row_values = [row_data[h] for h in headers]
            ws.append(row_values)
            
            current_row = ws.max_row
            status = row_data["_status"]
            
            if status == "mismatched_cost":
                for cell in ws[current_row]:
                    cell.fill = fill_yellow
            elif status == "unmatched_code":
                for cell in ws[current_row]:
                    cell.fill = fill_red
                    
        # Save File
        # yyyymmdd is dependent on selected CSV. Let's assume filename has date or use today's date?
        # User said: "yyyymmddは選択したcsvに依存"
        # Example CSV: 20251211未払合計.csv -> 20251211
        
        basename = os.path.basename(csv_path)
        match_date = re.search(r'\d{8}', basename)
        if match_date:
            date_str = match_date.group(0)
        else:
            from datetime import datetime
            date_str = datetime.now().strftime("%Y%m%d")
            
        output_filename = f"{date_str}マッチング済み.xlsx"
        output_path = os.path.join(os.path.dirname(csv_path), output_filename)
        
        wb.save(output_path)
        messagebox.showinfo("完了", f"マッチングが完了しました。\n保存先: {output_path}")

    except Exception as e:
        messagebox.showerror("エラー", f"予期せぬエラーが発生しました:\n{str(e)}")

if __name__ == "__main__":
    main()
